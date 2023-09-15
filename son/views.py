from django.shortcuts import render
from django.http import HttpResponse

from django.template import RequestContext
from openpyxl import load_workbook
# Create your views here.
from .models import *

from django.shortcuts import redirect


# Create your views here.


def home(request):
    query1 = "SELECT * from son_creance"   # Table creance in models  <==> Table son_creance in server phpmyadmin
    query2 = "SELECT * from son_commune"   # Table commune in models  <==> Table son_commune in server phpmyadmin
    
    Creance = creance.objects.raw(query1)  # Creance object but      Creance <=|=> creance 
    Commune = commune.objects.raw(query2)  # Commune object but      Commune <=|=> commune 
    print(Creance)
    context={                              # This context Going to to templet dash.html
          'Creance':Creance,
          'Commune' :Commune,
    }    

    return render(request,'son/dash.html',context)  # return page dash.html
#---------------------
def del_data(request):
    creance.objects.all().delete()
    return redirect(home)

#---------------------  
def import_xlsx(request):
    creance.objects.all().delete()
    if request.method=="POST":                         # Receiving a request from And Method Post

        file=request.POST.get('myfile')                # myfile is input file in from 
        mypath =request.POST.get('path')               # path is input file in from 
        workbook = load_workbook(filename=mypath+file) # Receiving mypath And file 
        sheet =workbook.active                         # Read file xlsx 
        rows  =sheet.iter_rows(min_row=2)              # Jump first line in xlsx
        for row in rows:                               # Read line by line in file 
            ref = row[0].value
            montant =row[1].value
            
            My_creance = creance()                     # Creat object My_creance from class creance()

            My_creance.ref = ref                       # Assign values to the corresponding in a file Xlsx
            My_creance.montant = montant               # Assign values to the corresponding in a file Xlsx

            My_creance.save()                          # save data and jump next line 

        return redirect(home)                          # return function home (call up function )


def show_creance(request):
    query1 = "SELECT * from son_creance"
    Creance = creance.objects.raw(query1)
    return render(request,'son/show_creance.html',{'Creance':Creance})


def function_1(request):
    function1.objects.all().delete()    #function1 is model

    # query = "SELECT id,ref,montant,ad_code from son_creance,son_commune where come_code=left(ref,5) "
    query = """SELECT  son_creance.id,ref, ad_code,montant FROM son_creance, son_commune 
               WHERE come_code = LEFT(ref, 5)   """

    f1_results=function1.objects.raw(query)
    
    for result in f1_results:
        f1_instance = function1()     
        f1_instance.id = result.id
        f1_instance.ref_f1 = result.ref
        f1_instance.montant_f1 = result.montant
        f1_instance.ad_code_f1 = result.ad_code
        f1_instance.save()

    f1 = function1.objects.raw("SELECT * from son_function1")
    return render(request,'son/function1.html',{'f1':f1})


#---------------------------------------------------------------

def function_2(request):
    function2.objects.all().delete()   #function2 is model 

    query = """SELECT son_function1.id,ref_f1,COUNT(*) As Nb , SUM(montant_f1) As montant_total ,ad_code_f1 FROM son_function1 
               GROUP BY ref_f1,ad_code_f1 """
    f2_results=function2.objects.raw(query)
    for result in f2_results:
        f2_instance = function2()
        f2_instance.id=result.id
        f2_instance.ref_f2=result.ref_f1
        f2_instance.Nb_Fact=result.Nb
        f2_instance.montant_f2=result.montant_total
        f2_instance.ad_code_f2=result.ad_code_f1
        f2_instance.save()

    f2 = function2.objects.raw("SELECT * from son_function2")

    return render(request,'son/function2.html',{'f2':f2})

#------------------------------------------------ --------------------------

def function_3(request):
    function3.objects.all().delete()   #function3is model 


    query = """SELECT son_function2.id,ad_code_f2 ,Nb_Fact,COUNT(Nb_Fact) As Nb,sum(montant_f2) as Montant from son_function2 
               GROUP BY ad_code_f2 , Nb_fact 
               order by NB ;
            """
    f3_results=function3.objects.raw(query)

    for result in f3_results:
        f3_instance = function3()
        f3_instance.id=result.id
        f3_instance.ad_code_f3=result.ad_code_f2
        f3_instance.Nb_Fact_all=result.Nb_Fact 
        f3_instance.Nb = result.Nb
        f3_instance.montant_f3 = result.Montant
        f3_instance.save()

    f3 = function3.objects.raw("SELECT * from son_function3")

    return render(request,'son/function3.html',{'f3':f3})
